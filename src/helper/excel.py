from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


class Excel:
    def __init__(self, filename):
        self.filename = filename
        self.project_name = None
        self.division = None
        self.employee_name = None
        self.employee_id = None
        self.period = None
        self.sheetname: str = ''
        self.sheetnames: list = []

    def read_exist(self):
        sheetname = self.sheetname
        try:
            if self.filename is not None:
                wbook = load_workbook(filename=self.filename)
                self.sheetnames = list(wbook.sheetnames)
                if sheetname in self.sheetnames: wsheet = wbook[sheetname]
                return wbook, wsheet
        except Exception as e:
            raise

    def fill_cell(self, values: list, row: int = 1, column: int = 1, orientation: str = 'vertical'):
        # DONE : Refactor this, make efficient with params
        wbook, wsheet = self.read_exist()
        try:
            for idx, value in enumerate(values):
                if orientation == 'vertical':
                    wsheet.cell(row=idx+row, column=column, value=value)
                if orientation == 'horizontal':
                    wsheet.cell(row=row, column=idx+column, value=value)

            wbook.save(self.filename)
        except Exception as e:
            raise

    def set_row(self, data: list, start_row):
        # DONE : Implement this (fill row based on row setting) Fill_cell
        pass
