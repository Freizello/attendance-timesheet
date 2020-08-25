import sys


def workbook():
    # TODO : Implement this workbook
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

    wb = Workbook()
    font = Font(
        name='Cambria',
        size=9
    )

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime

    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")


def tanggalmerah():
    # TODO : Implement this tanggalmerah

    from pytanggalmerah import TanggalMerah
    from datetime import datetime

    t = TanggalMerah()
    t.set_timezone("Asia/Jakarta")

    today = datetime.today()

    print(today.year, today.month, today.day)
    t.set_date(today.year, today.month, today.day)

    print(t.check())  # mengecek apakah tanggal merah, return boolean.
    print(t.is_holiday())  # mengecek apakah hari libur nasional, return boolean.
    print(t.is_sunday())  # mengecek apakah hari minggu, return booelan.
    print(t.get_event())  # mendapatkan event, return list.


if __name__ == '__main__':
    from src.helper.excel import Excel
    import src.config.constant as const

    ex = Excel(const.FILENAME_XLSX)
    ex.sheetname = "Timesheet"
    wbook, wsheet = ex.read_exist()

    # # print(wsheet['A1'].value, type(wsheet['A1'].value))
    #
    # cell_range = tuple(wsheet['A1':'C5'])
    # for row in cell_range:
    #     print(row)

    hcell = ['A1', 'A2', 'A3', 'A4', 'A5']
    vcell = ['C1', 'C2', 'C3', 'C4', 'C5']
    val = [
        const.PROJECT_NAME,
        const.UNIT,
        const.NAME,
        const.ID,
        const.PERIODE
    ]

    ex.set_header(header_cell=hcell, value_cell=vcell, value=val)